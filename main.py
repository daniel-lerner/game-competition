# -*- coding: utf-8 -*-
"""
Sistema de Pontuação para Competição de Jogos
Versão Orientada a Objetos
"""

import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
import os
from PIL import Image, ImageTk
import gspread
from oauth2client.service_account import ServiceAccountCredentials


class GameConfig:
    """Classe para gerenciar configurações dos jogos e jogadores"""
    
    def __init__(self):
        self.players = [
            'The_Lernos', 'Jujubex', 'The Rauls', 'Baumcy',
            'Camicaze', 'Pola destruidora', 'Mike Ty', 'Floydorc'
        ]
        
        self.games = [
            "Exploding Kittens", "Halli Galli", "Saco de Ossos", "Futebol de Moeda",
            "Ticket to Ride", "King of Tokyo", "Paper Town", "Abstratus",
            "Imagine", "Mille Fiori", "7 Wonders"
        ]


class UIManager:
    """Classe responsável pela interface gráfica"""
    
    def __init__(self):
        self.image_path = os.path.join(os.getcwd(), "Taverna_Lerner.webp")
    
    def create_selection_window(self, title, button_text, options):
        """Cria uma janela de seleção com dropdown"""
        selected_option = None
        
        def select_option():
            nonlocal selected_option
            selected_option = combo.get()
            window.quit()
        
        def on_closing():
            nonlocal selected_option
            selected_option = None
            window.quit()
        
        # Criar janela com tamanho adequado
        window = tk.Tk()
        window.title(title)
        window.geometry("800x600")  # 🔧 MUDANÇA: Tamanho fixo ao invés de maximizado
        window.protocol("WM_DELETE_WINDOW", on_closing)  # 🔧 MUDANÇA: Tratamento de fechamento
        
        # Centralizar janela
        window.update_idletasks()
        x = (window.winfo_screenwidth() // 2) - (800 // 2)
        y = (window.winfo_screenheight() // 2) - (600 // 2)
        window.geometry(f"800x600+{x}+{y}")
        
        # Label principal
        label = tk.Label(window, text=title, font=("Arial", 18), pady=20)
        label.pack()
        
        # Dropdown
        combo = ttk.Combobox(window, values=options, font=("Arial", 14), state="readonly")  # 🔧 MUDANÇA: readonly
        combo.pack(pady=10)
        combo.set(options[0])
        
        # Carregar imagem (com tratamento de erro)
        self._load_image_safe(window)  # 🔧 MUDANÇA: Método seguro para carregar imagem
        
        # Botão confirmar
        button = tk.Button(
            window, text=button_text, command=select_option, 
            font=("Arial", 14), pady=10, bg="#4CAF50", fg="white"  # 🔧 MUDANÇA: Cores do botão
        )
        button.pack(pady=20)
        
        window.mainloop()
        window.destroy()
        
        return selected_option
    
    def _load_image_safe(self, window):
        """Carrega imagem de forma segura"""
        try:
            if os.path.exists(self.image_path):
                img = Image.open(self.image_path)
                img = img.resize((400, 267), Image.Resampling.LANCZOS)  # 🔧 MUDANÇA: Tamanho menor
                img_tk = ImageTk.PhotoImage(img)
                
                label_img = tk.Label(window, image=img_tk)
                label_img.image = img_tk
                label_img.pack(pady=10)
            else:
                # 🔧 MUDANÇA: Placeholder quando imagem não existe
                placeholder = tk.Label(
                    window, text="🎲 Taverna dos Jogos 🎲", 
                    font=("Arial", 16), fg="#666", pady=20
                )
                placeholder.pack()
        except Exception as e:
            print(f"Aviso: Não foi possível carregar a imagem: {e}")
    
    def show_confirmation(self, player, game, num_players, position):
        """Mostra janela de confirmação"""
        message = f"{player} - {game} - {num_players} - {position}"
        
        # 🔧 MUDANÇA: Usar messagebox nativo do tkinter
        result = messagebox.askyesno(
            "Confirmação", 
            f"Confirmar pontuação?\n\n{message}",
            icon="question"
        )
        
        return result
    
    def show_final_message(self):
        """Mostra mensagem final"""
        # 🔧 MUDANÇA: Usar messagebox para mensagem final
        messagebox.showinfo(
            "Sucesso!", 
            "Pontos computados!\nVolte para a mesa e destrua seus inimigos!",
            icon="info"
        )
    
    def show_main_menu(self):
        """Mostra menu principal"""
        menu_options = [
            "Adicionar Pontuação",
            "Sair do Sistema"
        ]
        
        choice = self.create_selection_window(
            "Menu Principal - Sistema de Pontuação",
            "Selecionar",
            menu_options
        )
        
        return choice


class ScoreCalculator:
    """Classe responsável pelo cálculo de pontuação"""
    
    def calculate_points(self, game, num_players, position):
        """Calcula pontos baseado no jogo, número de jogadores e posição"""
        
        # 🔧 MUDANÇA: Organização mais clara da lógica de pontuação
        scoring_rules = {
            'Exploding Kittens': {
                "3 jogadores": {"1º lugar": 10},
                "4 jogadores": {"1º lugar": 12},
                "5 jogadores": {"1º lugar": 15}
            },
            'Halli Galli': {
                "4 jogadores": {"1º lugar": 10},
                "5 jogadores": {"1º lugar": 12},
                "6 jogadores": {"1º lugar": 15}
            },
            'Saco de Ossos': {
                "2 jogadores": {"1º lugar": 3}
            },
            'Futebol de Moeda': {
                "2 jogadores": {"1º lugar": 3}
            },
            'Ticket to Ride': {
                "3 jogadores": {"1º lugar": 60, "2º lugar": 30, "3º lugar": 20},
                "4 jogadores": {"1º lugar": 70, "2º lugar": 35, "3º lugar": 23},
                "5 jogadores": {"1º lugar": 80, "2º lugar": 40, "3º lugar": 26}
            },
            'King of Tokyo': {
                "4 jogadores": {"1º lugar": 60, "2º lugar": 30, "3º lugar": 20},
                "5 jogadores": {"1º lugar": 70, "2º lugar": 35, "3º lugar": 23},
                "6 jogadores": {"1º lugar": 80, "2º lugar": 40, "3º lugar": 26}
            },
            'Paper Town': {
                "3 jogadores": {"1º lugar": 40, "2º lugar": 20, "3º lugar": 0},
                "4 jogadores": {"1º lugar": 50, "2º lugar": 25, "3º lugar": 13}
            },
            'Abstratus': {
                "3 jogadores": {"1º lugar": 40, "2º lugar": 20, "3º lugar": 0},
                "4 jogadores": {"1º lugar": 50, "2º lugar": 25, "3º lugar": 13}
            },
            'Imagine': {
                "6 jogadores": {"1º lugar": 40, "2º lugar": 20, "3º lugar": 13},
                "7 jogadores": {"1º lugar": 50, "2º lugar": 25, "3º lugar": 16},
                "8 jogadores": {"1º lugar": 60, "2º lugar": 30, "3º lugar": 20}
            },
            'Mille Fiori': {
                "4 jogadores": {"1º lugar": 100, "2º lugar": 50, "3º lugar": 33}
            },
            '7 Wonders': {
                "5 jogadores": {"1º lugar": 70, "2º lugar": 35, "3º lugar": 23},
                "6 jogadores": {"1º lugar": 85, "2º lugar": 42, "3º lugar": 28},
                "7 jogadores": {"1º lugar": 100, "2º lugar": 50, "3º lugar": 33}
            }
        }
        
        try:
            return scoring_rules[game][num_players][position]
        except KeyError:
            print(f"Erro: Configuração não encontrada para {game}, {num_players}, {position}")
            return 0
    
    def get_player_options(self, game):
        """Retorna opções de número de jogadores para um jogo específico"""
        player_options = {
            'Exploding Kittens': ["3 jogadores", "4 jogadores", "5 jogadores"],
            'Halli Galli': ["4 jogadores", "5 jogadores", "6 jogadores"],
            'Saco de Ossos': ["2 jogadores"],
            'Futebol de Moeda': ["2 jogadores"],
            'Ticket to Ride': ["3 jogadores", "4 jogadores", "5 jogadores"],
            'King of Tokyo': ["4 jogadores", "5 jogadores", "6 jogadores"],
            'Paper Town': ["3 jogadores", "4 jogadores"],
            'Abstratus': ["3 jogadores", "4 jogadores"],
            'Imagine': ["6 jogadores", "7 jogadores", "8 jogadores"],
            'Mille Fiori': ["4 jogadores"],
            '7 Wonders': ["5 jogadores", "6 jogadores", "7 jogadores"]
        }
        return player_options.get(game, [])
    
    def get_position_options(self, game):
        """Retorna opções de posição para um jogo específico"""
        position_options = {
            'Exploding Kittens': ["1º lugar"],
            'Halli Galli': ["1º lugar"],
            'Saco de Ossos': ["1º lugar"],
            'Futebol de Moeda': ["1º lugar"],
            'Ticket to Ride': ["1º lugar", "2º lugar", "3º lugar"],
            'King of Tokyo': ["1º lugar", "2º lugar", "3º lugar"],
            'Paper Town': ["1º lugar", "2º lugar", "3º lugar"],
            'Abstratus': ["1º lugar", "2º lugar", "3º lugar"],
            'Imagine': ["1º lugar", "2º lugar", "3º lugar"],
            'Mille Fiori': ["1º lugar", "2º lugar", "3º lugar"],
            '7 Wonders': ["1º lugar", "2º lugar", "3º lugar"]
        }
        return position_options.get(game, [])


class GoogleSheetsManager:
    """Classe responsável pela integração com Google Sheets"""
    
    def __init__(self, credentials_file='credentials.json'):
        self.credentials_file = credentials_file
        self.client = None
        self.worksheet = None
        self._connect()
    
    def _connect(self):
        """Conecta com Google Sheets"""
        try:
            # 🔧 MUDANÇA: Tratamento de erro na conexão
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            credentials = ServiceAccountCredentials.from_json_keyfile_name(self.credentials_file, scope)
            self.client = gspread.authorize(credentials)
            self.worksheet = self.client.open('Tabela_de_pontos').sheet1
            print("✅ Conectado ao Google Sheets com sucesso!")
        except Exception as e:
            print(f"❌ Erro ao conectar com Google Sheets: {e}")
            self.client = None
            self.worksheet = None
    
    def update_score(self, player, game, points, games_list):
        """Atualiza pontuação na planilha"""
        if not self.worksheet:
            print("❌ Erro: Não conectado ao Google Sheets")
            return False
        
        try:
            # Buscar dados da planilha
            data = self.worksheet.get_all_records()
            
            # Encontrar linha do jogador
            player_row = None
            for index, row in enumerate(data):
                if row['Jogador'] == player:
                    player_row = index + 2  # +2 porque Google Sheets começa em 1 e primeira linha é cabeçalho
                    break
            
            if player_row is None:
                print(f"❌ Jogador {player} não encontrado na planilha")
                return False
            
            # Atualizar pontuação do jogo específico
            game_column = self.worksheet.find(game).col
            current_value = self.worksheet.cell(player_row, game_column).value
            current_value = int(current_value) if current_value and current_value.isdigit() else 0
            
            self.worksheet.update_cell(player_row, game_column, current_value + points)
            
            # Recalcular e atualizar total
            new_total = 0
            for game_name in games_list:
                try:
                    value = self.worksheet.cell(player_row, self.worksheet.find(game_name).col).value
                    new_total += int(value) if value and value.isdigit() else 0
                except:
                    continue
            
            total_column = self.worksheet.find('Total').col
            self.worksheet.update_cell(player_row, total_column, new_total)
            
            print(f"✅ Pontuação atualizada: {player} +{points} pontos em {game}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao atualizar planilha: {e}")
            return False


class GameScoreManager:
    """Classe principal que gerencia todo o sistema"""
    
    def __init__(self):
        self.config = GameConfig()
        self.ui = UIManager()
        self.calculator = ScoreCalculator()
        self.sheets_manager = GoogleSheetsManager()
    
    def collect_game_data(self):
        """Coleta dados do jogo do usuário"""
        try:
            # Selecionar jogador
            player = self.ui.create_selection_window(
                "Selecione o jogador", "Confirmar", self.config.players
            )
            if not player:
                return None
            
            # Selecionar jogo
            game = self.ui.create_selection_window(
                "Selecione o jogo", "Confirmar", self.config.games
            )
            if not game:
                return None
            
            # Selecionar número de jogadores
            player_options = self.calculator.get_player_options(game)
            num_players = self.ui.create_selection_window(
                "Selecione o número de jogadores que participaram", "Confirmar", player_options
            )
            if not num_players:
                return None
            
            # Selecionar posição
            position_options = self.calculator.get_position_options(game)
            position = self.ui.create_selection_window(
                "Selecione a posição final no jogo", "Confirmar", position_options
            )
            if not position:
                return None
            
            return player, game, num_players, position
            
        except Exception as e:
            print(f"❌ Erro ao coletar dados: {e}")
            return None
    
    def process_game_score(self):
        """Processa uma pontuação de jogo"""
        # Coletar dados
        game_data = self.collect_game_data()
        if not game_data:
            return False
        
        player, game, num_players, position = game_data
        
        # Confirmar dados
        if not self.ui.show_confirmation(player, game, num_players, position):
            print("❌ Operação cancelada pelo usuário")
            return False
        
        # Calcular pontos
        points = self.calculator.calculate_points(game, num_players, position)
        if points == 0:
            print("❌ Erro no cálculo de pontos")
            return False
        
        # Atualizar planilha
        success = self.sheets_manager.update_score(player, game, points, self.config.games)
        
        if success:
            self.ui.show_final_message()
            return True
        else:
            messagebox.showerror("Erro", "Falha ao atualizar a planilha!")
            return False
    
    def run(self):
        """Executa o sistema principal"""
        print("🎲 Sistema de Pontuação - Competição de Jogos")
        print("=" * 50)
        
        # 🔧 MUDANÇA: Menu principal ao invés de loop infinito
        while True:
            try:
                choice = self.ui.show_main_menu()
                
                if choice == "Adicionar Pontuação":
                    self.process_game_score()
                elif choice == "Sair do Sistema" or choice is None:
                    print("👋 Saindo do sistema...")
                    break
                else:
                    print("⚠️ Opção inválida")
                    
            except KeyboardInterrupt:
                print("\n👋 Sistema encerrado pelo usuário")
                break
            except Exception as e:
                print(f"❌ Erro inesperado: {e}")
                continue


# 🔧 MUDANÇA: Ponto de entrada principal
def main():
    """Função principal do programa"""
    try:
        game_manager = GameScoreManager()
        game_manager.run()
    except Exception as e:
        print(f"❌ Erro crítico: {e}")
        messagebox.showerror("Erro Crítico", f"Erro ao inicializar o sistema:\n{e}")


if __name__ == "__main__":
    main()